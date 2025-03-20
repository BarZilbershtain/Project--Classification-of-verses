import pandas as pd
import numpy as np
from anytree import Node, RenderTree
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

taamim_map = load_workbook('Teamim.xlsx')
sheet = taamim_map.active
data_dict = {}

# Define Teamim categories based on their role in the sentence structure
Ceaser = ['00', '92']  # Major separators
King = ['01', '65', '73', '80', '85']  # High-level connectors
Meshne = ['02', '03', '10', '88', '91']  # Secondary-level connectors
Shlise = ['14', '61', '62', '83', '84', '98']  # Third-level connectors

def build_the_tree(verse):
# Constructs a hierarchical tree structure based on Teamim categories.

    root = Node("root")  # Root node of the tree
    ceaser_nodes = []
    king_nodes = []
    meshne_nodes = []
    shlise_nodes = []

    for word in verse:
        if word not in [" ", "-"]:  # Ignore empty or non-relevant characters
            if word in Ceaser:
                # Create a node for major separators and attach previous king nodes
                ceaser_node = Node(word, parent=root)
                while king_nodes:
                    king_nodes[-1].parent = ceaser_node
                    king_nodes.pop()
                ceaser_nodes.append(ceaser_node)

            elif word in King:
                # Create a node for king-level connectors
                king_node = Node(word)
                while meshne_nodes:
                    meshne_nodes[-1].parent = king_node
                    meshne_nodes.pop()
                king_nodes.append(king_node)

            elif word in Meshne:
                # Create a node for secondary connectors
                meshne_node = Node(word)
                while shlise_nodes:
                    shlise_nodes[-1].parent = meshne_node
                    shlise_nodes.pop()
                meshne_nodes.append(meshne_node)
                
            elif word in Shlise:
                # Create a node for third-level connectors
                meshne_node = Node(word)
                while shlise_nodes:
                    shlise_nodes[-1].parent = meshne_node
                    shlise_nodes.pop()
                meshne_nodes.append(meshne_node)
                
            elif word in Shlise:
                # Append the word to the third-level list
                shlise_nodes.append(Node(word))

    # Attach remaining nodes to the root if they were not linked
    for king in king_nodes:
        king.parent = root
    for meshne in meshne_nodes:
        meshne.parent = root
    for shlise in shlise_nodes:
        shlise.parent = root
            
    return root

def teamimTree():
# Parses the Excel file and constructs trees for each verse.

    for row in sheet.iter_rows(values_only=True): 
        key = row[1]
        book = key[1:3]
        
        # Map book abbreviations to numerical identifiers
        book_map = {'gn': '0', 'ex': '1', 'lv': '2', 'nu': '3', 'dt': '4'}
        book = book_map.get(book)
        if book is None:
            continue  

        # Extract chapter and verse numbers
        chapter_pasuk = key[3:].split(':')
        chapter = chapter_pasuk[0]
        pasuk = chapter_pasuk[1].split(" ")[0]
        
        # Process the Teamim sequence
        value = row[2]
        value = value.split("/")
        value[0] = value[0].strip()  
        
        # Store the tree structure in the dictionary
        data_dict[f"{book}:{chapter}:{pasuk}"] = build_the_tree(value)

    return data_dict

def print_tree(tree, file):
    # Prints the hierarchical tree structure to a text file.
    
    for pre, _, node in RenderTree(tree):
        file.write(f"{pre}{node.name}\n")

# Generate trees for all verses
trees = teamimTree()

# Save the tree structures to a text file
with open("teamim-trees.txt", "w", encoding="utf-8") as file:
    for key, tree in trees.items():
        file.write(f"\nP: {key}\n")  
        print_tree(tree, file)
