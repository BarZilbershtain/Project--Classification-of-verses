import pandas as pd
from openpyxl import load_workbook

taamim_map = load_workbook('Teamim.xlsx')
sheet = taamim_map.active
data_dict = {}

# Define punctuation marks (Teamim) used to identify clause boundaries
Ceaser = ['00', '92']  # Clause-ending Taamim
King = ['01', '65', '73', '80', '85']  # Clause-continuing Teamim

def build_clauses(verse):
# Splits a verse into clauses based on Teamim

    clauses = []  
    current_clause = []  

    for word in verse:
        if word in Ceaser:  # If a word is a clause-ending Teamim
            if current_clause:
                current_clause.append(word)
                clauses.append(current_clause)
            current_clause = []  # Start a new clause
        elif word in King:  # If a word is a clause-continuing Teamim
            current_clause.append(word) 

    if current_clause:  
        clauses.append(current_clause)

    return clauses

def teamimTree():
    # Reads the Excel file and maps each verse to its corresponding Teamim clauses.

    for row in sheet.iter_rows(values_only=True): 
        key = row[1]  # Extract verse key
        book = key[1:3]  # Extract book identifier
        
        # Mapping biblical book abbreviations to numbers
        book_map = {'gn': '0', 'ex': '1', 'lv': '2', 'nu': '3', 'dt': '4'}
        
        book = book_map.get(book)
        if book is None:
            continue  
        
        # Extract chapter and verse number
        chapter_pasuk = key[3:len(key)].split(':')
        chapter = chapter_pasuk[0]
        pasuk = chapter_pasuk[1].split(" ")[0]
        
        # Process Teamim values
        value = row[2]
        value = str.split(value, "/")
        value[0] = value[0].removeprefix(" ")
        
        # Store the parsed clauses in the dictionary
        data_dict[f"{book}:{chapter}:{pasuk}"] = build_clauses(value)

    return data_dict

def print_clauses_to_txt():
    # Writes the extracted clauses into a text file

    with open('teamim-clauses.txt', 'w', encoding='utf-8') as f:
        for key, clauses in data_dict.items():
            f.write(f"verse: {key}:\n")
            for i, clause in enumerate(clauses):
                f.write(f"  {i}. {' '.join(clause)}\n")
            f.write("\n")

# Generate the clause mapping and save the output
data_dict = teamimTree()
print_clauses_to_txt()
